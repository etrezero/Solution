# 필요한 패키지 임포트
from dash import Dash, html, dash_table, Input, Output, dcc 
from dash.dash_table.Format import Format, Scheme
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import warnings
from tqdm import tqdm
import os
from openpyxl import Workbook
from flask import Flask
from openpyxl.utils import get_column_letter, column_index_from_string


# Flask 서버 생성
server = Flask(__name__)


# 앱 초기화
app = Dash(__name__, suppress_callback_exceptions=True, server=server)
app.title = 'TDF 메인'  # 브라우저 탭 제목 설정


# 경고 메시지 무시
warnings.filterwarnings("ignore", category=UserWarning)



# 모니터링 ---------> temp
path_TDF = r'C:/Covenant/data/0.TDF_모니터링.xlsx'  # 이 경로를 정확히 지정해야 합니다.
sheet_RANK = 'RANK'  # '요약.xlsx' 파일에 있는 시트 이름

path_Temp_TDF = r'C:/Covenant/data/Temp_TDF.xlsx'
sheet_temp = 'RANK'



# # 지정된 시트를 읽어와서 데이터프레임으로 저장
# df_RANK = pd.read_excel(path_TDF, sheet_name=sheet_RANK)

# # 파일이 없는 경우 새 Workbook 생성
# if not os.path.exists(path_Temp_TDF):
#     wb = Workbook()
#     wb.save(path_Temp_TDF)
#     print(f"새 파일 '{path_Temp_TDF}' 생성됨.")

# # DataFrame을 엑셀 시트로 저장
# with pd.ExcelWriter(path_Temp_TDF, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#     df_RANK.to_excel(writer, sheet_name=sheet_temp, index=False)

# print(f"'{sheet_RANK}' 시트를 '{sheet_temp}' 시트로 저장했습니다.")



# 공통 컬럼 스타일 설정
column_styles = {
    'percentage': {
        'indexes': [2, 5, 8, 11, 14, 17, 20, 23], 
        'style': {'type': 'numeric', 
                  'format': Format(precision=1, scheme=Scheme.percentage), 
                  'textAlign': 'center'
                },
                
        },

    'left': {
        'indexes': [1, 4, 7, 10, 13, 16, 19, 22], 
        'style': {'textAlign': 'left'},
        'width': '3%',
        },

    'number': {
        'indexes': [0, 3, 6, 9, 12, 15, 18, 21], 
        'style': {'textAlign': 'center'}
    },
}


# 테이블 스타일 및 컬럼 스타일 설정
table_cell = {
    'table1': {'range': 'C3:Z26', 
               'styles': column_styles, 
               'reverse': False},
    'table2': {'range': 'C34:Z57', 
               'styles': column_styles, 
               'reverse': True},
    'table3': {'range': 'C63:Z86', 
               'styles': {
                'percentage': {
                    'indexes': [2, 5, 8, 11, 14, 17, 20, 23], 
                    'style': {'type': 'numeric', 
                            'format': Format(precision=1, scheme=Scheme.decimal), 
                            'textAlign': 'center'}
                },
                'left': {
                    'indexes': [1, 4, 7, 10, 13, 16, 19, 22], 
                    'style': {'textAlign': 'left'},
                    'width': '3%',
                },
                'number': {
                    'indexes': [0, 3, 6, 9, 12, 15, 18, 21], 
                    'style': {'textAlign': 'center'}
                },
                'reverse': False
            }}
}



# E2 셀에서 term 값 읽어오기
term = load_workbook(path_Temp_TDF)[sheet_temp]['E2'].value

# 테이블 타이틀 설정
table_titles = {
    'table1': f'수익률 {term}', 
    'table2': f'변동성 {term}', 
    'table3': f'위험대비수익률 {term}'
}


# 전체 테이블 및 헤더 스타일
table_style = {
    'margin': '0', 
    'margin-top': '-1%', 
    'margin-bottom': '1%', 
    'width': '100%', 
    'textAlign': 'center', 
    'z-index': 3
}

table_header_style = {
    'fontWeight': 'bold', 
    'backgroundColor': '#3762AF', 
    'textAlign': 'center', 
    'color': 'white'
}

# 데이터프레임을 생성하는 함수
def create_data_table(path, sheet_name, cell_range_key):
    cell_range_str = table_cell[cell_range_key]['range']
    start_cell, end_cell = cell_range_str.split(':')
    start_col, start_row = start_cell[:1], int(start_cell[1:])
    end_col, end_row = end_cell[:1], int(end_cell[1:])
    
    df = pd.read_excel(path, sheet_name=sheet_name, engine='openpyxl', usecols=f"{start_col}:{end_col}", skiprows=start_row - 1, nrows=end_row - start_row + 1, header=0)
    df.columns = df.columns.astype(str)

    columns = []
    style_cell_conditional = []

    # 컬럼 설정
    for idx, col in enumerate(df.columns):
        column_info = {"name": col, "id": col}
        
        # "Unnamed"으로 시작하는 컬럼 이름은 빈 문자열로 대체
        col_name = "" if col.startswith("Unnamed") else col
        column_info = {"name": col_name, "id": col}
        
        # 타입 및 포맷 설정
        if 'percentage' in table_cell[cell_range_key]['styles'] and idx in table_cell[cell_range_key]['styles']['percentage']['indexes']:
            column_info.update({
                "type": "numeric",
                "format": table_cell[cell_range_key]['styles']['percentage']['style']['format']
            })

        columns.append(column_info)

    # 스타일 조건 설정
    for style_name, style_info in table_cell[cell_range_key]['styles'].items():
        if isinstance(style_info, dict):
            for idx in style_info.get('indexes', []):
                if idx < len(df.columns):
                    col = df.columns[idx]
                    style_cell_conditional.append({
                        'if': {'column_id': col},
                        'textAlign': style_info.get('style', {}).get('textAlign', 'center'),
                        'format': style_info.get('style', {}).get('format', None)
                    })

    return dash_table.DataTable(
        id='table',
        columns=columns,
        data=df.to_dict('records'),
        style_table=table_style,
        style_header=table_header_style,
        style_cell_conditional=style_cell_conditional[::-1] if table_cell[cell_range_key].get('reverse', False) else style_cell_conditional
    )


# 레이아웃 생성
app.layout = html.Div([
    html.Div([
        html.H3(table_titles[key]),
        create_data_table(
            path=path_Temp_TDF,
            sheet_name=sheet_temp,
            cell_range_key=key
        )
    ], className='table') for key in table_cell.keys()
])


# 앱 실행
if __name__ == '__main__':
    app.run_server(debug=True, host='0.0.0.0')
