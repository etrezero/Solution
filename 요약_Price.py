import FinanceDataReader as fdr
from datetime import datetime, timedelta
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle

# 엑셀 파일 경로
save_path = 'C:/Covenant/data/요약.xlsx'

# 오늘 날짜를 변수에 할당
today = datetime.now()
today_str = today.strftime('%Y-%m-%d')
print(today_str)

# 시작 날짜 설정
Start = today - timedelta(days=367)
End = today - timedelta(days=0)



def fetch_and_save_data(sheet_name, 종목코드_list):
    data_frames = []
    for 종목코드 in tqdm(종목코드_list, desc=f'Fetching data for {sheet_name}'):
        try:
            df = fdr.DataReader(종목코드, Start, End)['Close'].rename(종목코드)
            df.index = df.index.strftime('%Y-%m-%d')
            data_frames.append(df)
        except Exception as e:
            print(f"Error fetching data for 종목코드 '{종목코드}': {e}")
    combined_df = pd.concat(data_frames, axis=1)
    with pd.ExcelWriter(save_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        combined_df.to_excel(writer, sheet_name=sheet_name, index=True)

    wb = load_workbook(save_path)/
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)

    print(f'데이터가 {save_path}의 {sheet_name} 시트에 저장되었습니다.')

# List1 데이터 가져오기
df_list1 = pd.read_excel(save_path, sheet_name='List1', usecols=[0], names=['종목코드'])
종목코드_list1 = df_list1['종목코드'].astype(str).tolist()
fetch_and_save_data('Price1', 종목코드_list1)


