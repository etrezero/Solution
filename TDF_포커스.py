# 필요한 패키지 임포트
from dash import Dash, dcc, html, dash_table, Input, Output
from dash.dash_table.Format import Format, Scheme
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import warnings
from tqdm import tqdm
import os
from openpyxl import Workbook


#[Dash 그래프 요소]
# 그래프 유형: Line (선), Bar (막대), Scatter (산점도), Pie (원형), Heatmap (히트맵), Boxplot (상자그림), Histogram (히스토그램), 3D 그래프 등 다양한 그래프 유형을 제공합니다.
# 축 설정: X축과 Y축의 레이블, 범위, 타입 (선형, 로그 등), 틱 마크 설정 등.
# 레이아웃 설정: 그래프의 제목, 글꼴 스타일, 그래프의 크기, 배경 색상, 레이아웃 마진 설정 등.
# 마커와 라인 스타일: 산점도의 점 크기와 색상, 선 그래프의 선 굵기와 스타일 등.
# 범례 (Legend): 범례의 위치와 스타일 설정.
# 호버 텍스트: 데이터 포인트에 마우스를 가져갔을 때 표시되는 정보.
# 애니메이션: 데이터 변화에 따른 그래프의 동적 변화를 표현.
# 콜백 및 이벤트 처리: 사용자 상호작용에 따른 그래프의 동적 업데이트.
# 서브플롯 및 다중 축: 하나의 그래프 안에 여러 개의 서브플롯 또는 다중 축을 구현.
# 툴팁: 사용자가 그래프의 특정 부분에 마우스를 올렸을 때 추가 정보를 제공.
# 배경 색상: 그래프 전체 배경의 색상을 설정할 수 있습니다. 이는 그래프 데이터를 강조하거나 특정 테마에 맞게 그래프를 조정하는 데 유용합니다.
# 서브플롯 배경: 다중 그래프 또는 서브플롯의 배경 색상을 별도로 설정할 수 있습니다. 각 서브플롯이 서로 구분되도록 시각적 구분을 제공합니다.
# 플롯 영역 배경: X축과 Y축 사이의 주요 그래프 영역의 배경을 설정할 수 있습니다. 이 영역은 데이터가 표시되는 주요 부분입니다.
# 그리드 라인: 그래프의 배경에 그리드 라인을 추가하여 데이터 포인트의 위치를 더 쉽게 식별할 수 있습니다. 그리드 라인의 색상, 스타일, 두께를 조정할 수 있습니다.
# 그림자 효과: 그래프에 그림자 효과를 추가하여 시각적 깊이와 입체감을 줄 수 있습니다.
# 테두리: 그래프의 외곽에 테두리를 추가하여 그래프를 나머지 페이지 또는 대시보드 요소와 구분할 수 있습니다.
# 불투명도: 그래프의 배경 불투명도를 조정하여 배경 이미지나 색상을 더 강조하거나 더 희미하게 할 수 있습니다.
# 배경 이미지: 배경으로 이미지를 설정하여 그래프에 추가적인 맥락이나 시각적 효과를 제공할 수 있습니다.
# 마진 및 패딩: 그래프 주변의 여백을 조정하여 배경 요소와의 관계를 설정할 수 있습니다.


# 경고 메시지 무시
warnings.filterwarnings("ignore", category=UserWarning)



# 모니터링 ---------> temp
path_TDF = r'C:/Covenant/data/0.TDF_모니터링.xlsx'  # 이 경로를 정확히 지정해야 합니다.
sheet_요약 = '요약'  # '요약.xlsx' 파일에 있는 시트 이름
path_Temp_TDF = r'C:/Covenant/data/Temp_TDF.xlsx'
sheet_temp = 'temp'

# 지정된 시트를 읽어와서 데이터프레임으로 저장
df_요약 = pd.read_excel(path_TDF, sheet_name=sheet_요약)

# 파일이 없는 경우 새 Workbook 생성
if not os.path.exists(path_Temp_TDF):
    wb = Workbook()
    wb.save(path_Temp_TDF)
    print(f"새 파일 '{path_Temp_TDF}' 생성됨.")

# DataFrame을 엑셀 시트로 저장
with pd.ExcelWriter(path_Temp_TDF, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_요약.to_excel(writer, sheet_name=sheet_temp, index=False)

print(f"'{sheet_요약}' 시트를 '{sheet_temp}' 시트로 저장했습니다.")




# 앱 초기화
app = Dash(__name__)

# 엑셀 파일 내 특정 셀 범위에서 데이터를 읽어오는 함수
def read_data_from_excel(path_Temp_TDF, sheet_temp, cell_range):
    wb = openpyxl.load_workbook(path_Temp_TDF, data_only=True)
    sheet = wb[sheet_temp]
    data = []
    for row in sheet[cell_range]:
        data.append([cell.value for cell in row])
    df = pd.DataFrame(data[1:], columns=data[0])
    return df

    # df.columns = [str(col) for col in df.columns]


# 테이블의 데이터 불러오기
df_table1 = read_data_from_excel(path_Temp_TDF, sheet_temp, 'A3:G11')
df_table2 = read_data_from_excel(path_Temp_TDF, sheet_temp, 'I3:O11')
df_table3 = read_data_from_excel(path_Temp_TDF, sheet_temp, 'Q3:U11')
df_table4 = read_data_from_excel(path_Temp_TDF, sheet_temp, 'W3:AA12')
df_table5 = read_data_from_excel(path_Temp_TDF, sheet_temp, 'AC3:AK9')
df_table6 = read_data_from_excel(path_Temp_TDF, sheet_temp, 'AW43:BE58')



# 데이터 테이블을 생성하는 함수
def create_data_table(path_Temp_TDF, sheet_temp, cell_range, percent_columns, font_size=None, header_style=None):
    df = read_data_from_excel(path_Temp_TDF, sheet_temp, cell_range)
    
    # 열의 너비 계산
    num_columns = len(df.columns)
    column_width = "{}%".format(100 / num_columns)
            
      
    # 첫 번째 열을 텍스트로 표시
    columns = [{"name": str(i), "id": str(i), "type": "text" if i == df.columns[0] else "numeric"} for i in df.columns]

    # 모든 열을 numeric으로 설정 (첫 번째 열은 텍스트로 설정됨)
    for column in columns[1:]:
        column['type'] = 'numeric'
        column['format'] = Format(precision=0, scheme=Scheme.fixed)  # 모든 숫자 열의 소수 자릿수를 0으로 설정

    # 퍼센트 컬럼으로 지정한 열만 소수점 2자리까지 보이도록 설정
    if percent_columns:
        for idx in percent_columns:
            adjusted_idx = idx - 1
            if 0 <= adjusted_idx < len(columns):
                columns[adjusted_idx]['format'] = Format(precision=1, scheme=Scheme.percentage)

    cell_style = {
        'width': column_width,
        'minWidth': column_width,
        'maxWidth': column_width,
        'whiteSpace': 'normal',
        'textAlign': 'center',
        'verticalAlign': 'middle',
        'fontSize': '12px',
        # 'fontWeight': 'bold',
        'overflow': 'hidden',
        'textOverflow': 'ellipsis'
    }
    
    # 헤더 스타일 변경
    style_header = {
        'color': 'White', 
        'fontweight': 'Bold', 
        'background-color': '#3762AF', #'darkblue'
        # 'background-color': '#4BACC6', #연한 비취색
    }

    
    return dash_table.DataTable(
        data=df.to_dict('records'),
        columns=columns,
        page_size=9,  #행 개수 넘어가면 페이지 넘어가
        style_cell=cell_style,  # 테이블 셀 텍스트 정렬을 중앙으로 설정
        style_header=style_header,
    )



# 마지막 행이 없는 데이터 - 그래프용
df_G1 = df_table1.iloc[:-1]
df_G2 = df_table2.iloc[:-1]
df_G3 = df_table3.iloc[:-1]
df_G4 = df_table4.iloc[:-1]
df_G5 = df_table5.iloc[:-1]
df_G6 = df_table6.iloc[:-1]

# 테이블 셀 범위 정의
table_cell = {
    'table1': 'A3:G11',
    'table2': 'I3:O11',
    'table3': 'Q3:U12',
    'table4': 'W3:AA12',
    'table5': 'AC3:AK9',
    'table6': 'AW43:BE58'
}

percent_column = {
    '%table1' : [2, 3, 4, 5, 6, 7],
    '%table2' : [2, 3, 4, 5, 6, 7],
    '%table3' : [3, 5],
    '%table4' : [3, 5],
    '%table5' : [2, 3, 4, 5, 6, 7, 8, 9],
    '%table6' : [2, 3, 4, 5, 6, 7, 8, 9],
}

# 그리드 아이템 CSS <A구역 : B구역>
grid_style = {
    'display': 'inline-block',  # 인라인 블록 요소로 배치
    'width': '49%',  # 화면 절반을 차지
    'margin': '0 auto',
    'verticalAlign': 'top',  # 상단 정렬
    # 'padding' :200,
    # 'paddingLeft': '100px', 
    # 'border': '1px solid black',  # 바더 라인 추가
}

# 테이블 스타일
table_style = {
    'margin': '0 auto',
    'margin-top': '7%',
    'width': '60%',
    'textAlign': 'center',
    # 'fontsize' : '30px',
    'z-index': 3,  # 이 요소가 다른 요소 위에 위치함
    # 'padding' :300,
    # 'leftpadding' :300,
    # 'rightpadding' :300,
}

# 그래프 형식 (Line, Bar, Dot, Pie) 목록
graph_types = ['Line', 'Bar', 'Dot', 'Pie','Heatmap','Histogram']

# 그래프 컨테이너 스타일
graph_container_style = {
    'position': 'relative',  # 상대적 위치 설정
    'margin': '0 auto',
    'width': '90%',  # 컨테이너 너비
    'z-index': 1,  # 이 요소가 다른 요소 위에 위치함
    # 'border': '1px solid black',  # 바더 라인 추가
}

# 그래프 스타일 설정
graph_style = {
    'margin': '0 auto',
    'margin-top': '0%',
    'width': '85%',
    'z-index': 1,  # 이 요소가 다른 요소 위에 위치함
}


# 드롭다운 컨테이너 스타일
dropdown_container_style = {
    'position': 'absolute',  # 절대적 위치 설정
    'width': '15%',  # 각 드롭다운의 너비
    'right': '0%',  # 오른쪽에서 5% 떨어진 곳에 위치
    'top': '70%',  # 상단에서 50% 위치
    'transform': 'translateY(-50%)',  # Y축으로 -50% 이동하여 중앙 정렬
    'border': '0px solid black',  # 바더 라인 추가
    'z-index': 3,  # 이 요소가 다른 요소 위에 위치함
}

# 개별 드롭다운 스타일
dropdown_style = {
    'width': '100%',  # 각 드롭다운의 너비
    'marginRight': '1%',  # 드롭다운 사이의 간격
    'textAlign': 'center',
}


# 앱 레이아웃 설정
app.layout = html.Div([
    # html.Div(children='한국투자 TDF ETF 포커스 모니터링', 
    #          style={
    #              'fontSize': 25, 
    #              'textAlign': 'center', 
    #              'position': 'sticky', 
    #              'top': '0', 
    #              'zIndex': '100',
    #              'background-color': 'white'
    #          }),
    
    # 테이블 1 
    html.Div([
        # 테이블 컨테이너
        html.Div([
            html.H3('수익률'),
            create_data_table(
                path_Temp_TDF, 
                sheet_temp, 
                table_cell['table1'],
                percent_column['%table1'], 
                12, 
                {'backgroundColor': 'blue', 'textAlign': 'center'}
            )
        ], className='table', style=table_style),

        # 그래프 컨테이너
        html.Div([
            dcc.Graph(id='line-graph-table1', style=graph_style),

            # 드롭다운 컨테이너
            html.Div([
                dcc.Dropdown(
                    id='yaxis-column-table1',
                    options=[{'label': i, 'value': i} for i in df_table1.columns[1:]],
                    value=df_table1.columns[5],
                    style=dropdown_style
                ),
                dcc.Dropdown(
                    id='graph-type-table1',
                    options=[{'label': i, 'value': i} for i in graph_types],
                    value='Dot',
                    style=dropdown_style
                )
            ], className='dropdown', style=dropdown_container_style),
        ], className='graph', style=graph_container_style),
    ], className='grid', style=grid_style),


    # 테이블 2
    html.Div([
        # 테이블 컨테이너
        html.Div([
            html.H3('변동성'),
            create_data_table(
                path_Temp_TDF, 
                sheet_temp, 
                table_cell['table2'],
                percent_column['%table2'], 
                12, 
                {'backgroundColor': 'blue', 'textAlign': 'center'}
            )
        ], className='table', style=table_style),

        # 그래프 컨테이너
        html.Div([
            dcc.Graph(id='line-graph-table2', style=graph_style),

            # 드롭다운 컨테이너
            html.Div([
                dcc.Dropdown(
                    id='yaxis-column-table2',
                    options=[{'label': i, 'value': i} for i in df_table2.columns[1:]],
                    value=df_table2.columns[5],
                    style=dropdown_style
                ),
                dcc.Dropdown(
                    id='graph-type-table2',
                    options=[{'label': i, 'value': i} for i in graph_types],
                    value='Dot',
                    style=dropdown_style
                )
            ], className='dropdown', style=dropdown_container_style),
        ], className='graph', style=graph_container_style),
    ], className='grid', style=grid_style),


    # 테이블 3
    html.Div([
        # 테이블 컨테이너
        html.Div([
            html.H3('설정액(시장전체)'),
            create_data_table(
                path_Temp_TDF, 
                sheet_temp, 
                table_cell['table3'],
                percent_column['%table3'], 
                12, 
                {'backgroundColor': 'blue', 'textAlign': 'center'}
            )
        ], className='table', style=table_style),

        # 그래프 컨테이너
        html.Div([
            dcc.Graph(id='line-graph-table3', style=graph_style),

            # 드롭다운 컨테이너
            html.Div([
                dcc.Dropdown(
                    id='yaxis-column-table3',
                    options=[{'label': i, 'value': i} for i in df_table3.columns[1:]],
                    value=df_table3.columns[1],
                    style=dropdown_style
                ),
                dcc.Dropdown(
                    id='graph-type-table3',
                    options=[{'label': i, 'value': i} for i in graph_types],
                    value='Pie',
                    style=dropdown_style
                )
            ], className='dropdown', style=dropdown_container_style),
        ], className='graph', style=graph_container_style),
    ], className='grid', style=grid_style),



    # 테이블 4 
    html.Div([
        # 테이블 컨테이너
        html.Div([
            html.H3('설정액(포커스)'),
            create_data_table(
                path_Temp_TDF, 
                sheet_temp, 
                table_cell['table4'],
                percent_column['%table4'], 
                12, 
                {'backgroundColor': 'blue', 'textAlign': 'center'}
            )
        ], className='table', style=table_style),

        # 그래프 컨테이너
        html.Div([
            dcc.Graph(id='line-graph-table4', style=graph_style),

            # 드롭다운 컨테이너
            html.Div([
                dcc.Dropdown(
                    id='yaxis-column-table4',
                    options=[{'label': i, 'value': i} for i in df_table4.columns[1:]],
                    value=df_table4.columns[1],
                    style=dropdown_style
                ),
                dcc.Dropdown(
                    id='graph-type-table4',
                    options=[{'label': i, 'value': i} for i in graph_types],
                    value='Pie',
                    style=dropdown_style
                )
            ], className='dropdown', style=dropdown_container_style),
        ], className='graph', style=graph_container_style),
    ], className='grid', style=grid_style),


    # 테이블 5
    html.Div([
        # 테이블 컨테이너
        html.Div([
            html.H3('자산배분'),
            create_data_table(
                path_Temp_TDF, 
                sheet_temp, 
                table_cell['table5'],
                percent_column['%table5'], 
                12, 
            )
        ], className='table', style=table_style),

        # 그래프 컨테이너
        html.Div([
            dcc.Graph(id='line-graph-table5', style=graph_style),

            # 드롭다운 컨테이너
            html.Div([
                dcc.Dropdown(
                    id='yaxis-column-table5',
                    options=[{'label': i, 'value': i} for i in df_table5.columns[1:]],
                    value=df_table5.columns[1],
                    style=dropdown_style
                ),
                dcc.Dropdown(
                    id='graph-type-table5',
                    options=[{'label': i, 'value': i} for i in graph_types],
                    value='Pie',
                    style=dropdown_style
                )
            ], className='dropdown', style=dropdown_container_style),
        ], className='graph', style=graph_container_style),
    ], className='grid', style=grid_style),


    # 테이블 6
    html.Div([
        # 테이블 컨테이너
        html.Div([
            html.H3('투자비중'),
            create_data_table(
                path_Temp_TDF, 
                sheet_temp, 
                table_cell['table6'],
                percent_column['%table6'], 
                12, 
                {'backgroundColor': 'blue', 'textAlign': 'center'}
            )
        ], className='table', style=table_style),

        # 그래프 컨테이너
        html.Div([
            dcc.Graph(id='line-graph-table6', style=graph_style),

            # 드롭다운 컨테이너
            html.Div([
                dcc.Dropdown(
                    id='yaxis-column-table6',
                    options=[{'label': i, 'value': i} for i in df_table6.columns[1:]],
                    value=df_table6.columns[6],
                    style=dropdown_style
                ),
                dcc.Dropdown(
                    id='graph-type-table6',
                    options=[{'label': i, 'value': i} for i in graph_types],
                    value='Bar',
                    style=dropdown_style
                )
            ], className='dropdown', style=dropdown_container_style),
        ], className='graph', style=graph_container_style),
    ], className='grid', style=grid_style),


])


# 콜백 함수 정의
@app.callback(
    Output('line-graph-table1', 'figure'),
    [Input('yaxis-column-table1', 'value'),
     Input('graph-type-table1', 'value')]
)
def update_graph_table1(yaxis_column_name, graph_type):
    if graph_type == 'Dot':
        # Dot 그래프 생성
        fig = go.Figure(data=go.Scatter(
            x=df_table1[df_table1.columns[0]].astype(str),
            y=df_table1[yaxis_column_name],
            mode='markers',  # 마커 모드 사용
            marker=dict(
                size=15,  # 데이터 마크 크기 조절 (원하는 크기로 설정)
                opacity=0.8,
                line=dict(width=2, color='DarkSlateGrey')
            )
        ))
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정
        fig.update_xaxes(tickformat="Value", tickangle=-45, tickmode='auto', nticks=12) # 여기서 nticks를 조정하여 간격 변경
        
    elif graph_type == 'Bar':
        fig = px.bar(df_table1, x=df_table1.columns[0], y=yaxis_column_name)
    elif graph_type == 'Pie':
        fig = px.pie(df_table1, names=df_table1.columns[0], values=yaxis_column_name)
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정
    else:
        fig = px.line(df_table1, x=df_table1.columns[0], y=yaxis_column_name)
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정

            # 그래프의 배경을 투명하게 설정
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        # plot_bgcolor='rgba(0,0,0,0)',
    )
    return fig





@app.callback(
    Output('line-graph-table2', 'figure'),
    [Input('yaxis-column-table2', 'value'),
     Input('graph-type-table2', 'value')]
)

def update_graph_table2(yaxis_column_name, graph_type):
    if graph_type == 'Dot':
        # Dot 그래프 생성
        fig = go.Figure(data=go.Scatter(
            x=df_table2[df_table2.columns[0]].astype(str),
            y=df_table2[yaxis_column_name],
            mode='markers',  # 마커 모드 사용
            marker=dict(
                size=15,  # 데이터 마크 크기 조절 (원하는 크기로 설정)
                opacity=0.8,
                line=dict(width=2, color='DarkSlateGrey')
            )
        ))
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정
        fig.update_xaxes(tickformat="Value", tickangle=-45, tickmode='auto', nticks=12) # 여기서 nticks를 조정하여 간격 변경
        
    elif graph_type == 'Bar':
        fig = px.bar(df_table2, x=df_table2.columns[0], y=yaxis_column_name)
    elif graph_type == 'Pie':
        fig = px.pie(df_table2, names=df_table2.columns[0], values=yaxis_column_name)
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정
    else:
        fig = px.line(df_table2, x=df_table2.columns[0], y=yaxis_column_name)
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정

    # 그래프의 배경을 투명하게 설정
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        # plot_bgcolor='rgba(0,0,0,0)'
    )

    return fig



@app.callback(
    Output('line-graph-table3', 'figure'),
    [Input('yaxis-column-table3', 'value'),
     Input('graph-type-table3', 'value')]
)

def update_graph_table3(yaxis_column_name, graph_type):
    if graph_type == 'Dot':
        # Dot 그래프 생성
        fig = go.Figure(data=go.Scatter(
            x=df_G3[df_G3.columns[0]].astype(str),
            y=df_G3[yaxis_column_name],
            mode='markers',  # 마커 모드 사용
            marker=dict(
                size=15,  # 데이터 마크 크기 조절 (원하는 크기로 설정)
                opacity=0.8,
                line=dict(width=2, color='DarkSlateGrey')
            )
        ))
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정
        fig.update_xaxes(tickformat="Value", tickangle=-45, tickmode='auto', nticks=12) # 여기서 nticks를 조정하여 간격 변경
        
    elif graph_type == 'Bar':
        fig = px.bar(df_G3, x=df_G3.columns[0], y=yaxis_column_name)
    elif graph_type == 'Pie':
        fig = px.pie(df_G3, names=df_G3.columns[0], values=yaxis_column_name)
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정
    else:
        fig = px.line(df_G3, x=df_G3.columns[0], y=yaxis_column_name)
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정

    # 그래프의 배경을 투명하게 설정
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        # plot_bgcolor='rgba(0,0,0,0)',
    )

    return fig


@app.callback(
    Output('line-graph-table4', 'figure'),
    [Input('yaxis-column-table4', 'value'),
     Input('graph-type-table4', 'value')]
)

def update_graph_table4(yaxis_column_name, graph_type):
    if graph_type == 'Dot':
        # Dot 그래프 생성
        fig = go.Figure(data=go.Scatter(
            x=df_G4[df_G4.columns[0]].astype(str),
            y=df_G4[yaxis_column_name],
            mode='markers',  # 마커 모드 사용
            marker=dict(
                size=15,  # 데이터 마크 크기 조절 (원하는 크기로 설정)
                opacity=0.8,
                line=dict(width=2, color='DarkSlateGrey')
            )
        ))
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정
        fig.update_xaxes(tickformat="Value", tickangle=-45, tickmode='auto', nticks=12) # 여기서 nticks를 조정하여 간격 변경
        
    elif graph_type == 'Bar':
        fig = px.bar(df_G4, x=df_G4.columns[0], y=yaxis_column_name)
    elif graph_type == 'Pie':
        fig = px.pie(df_G4, names=df_G4.columns[0], values=yaxis_column_name)
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정
    else:
        fig = px.line(df_G4, x=df_G4.columns[0], y=yaxis_column_name)
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정

    # 그래프의 배경을 투명하게 설정
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        # plot_bgcolor='rgba(0,0,0,0)',
    )

    return fig


@app.callback(
    Output('line-graph-table5', 'figure'),
    [Input('yaxis-column-table5', 'value'),
     Input('graph-type-table5', 'value')]
)

def update_graph_table5(yaxis_column_name, graph_type):
    if graph_type == 'Dot':
        # Dot 그래프 생성
        fig = go.Figure(data=go.Scatter(
            x=df_table5[df_table5.columns[0]].astype(str),
            y=df_table5[yaxis_column_name],
            mode='markers',  # 마커 모드 사용
            marker=dict(
                size=15,  # 데이터 마크 크기 조절 (원하는 크기로 설정)
                opacity=0.8,
                line=dict(width=2, color='DarkSlateGrey')
            )
        ))
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정
        fig.update_xaxes(tickformat="Value", tickangle=-45, tickmode='auto', nticks=12) # 여기서 nticks를 조정하여 간격 변경
        
    elif graph_type == 'Bar':
        fig = px.bar(df_table5, x=df_table5.columns[0], y=yaxis_column_name)
    elif graph_type == 'Pie':
        fig = px.pie(df_table5, names=df_table5.columns[0], values=yaxis_column_name)
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정
    else:
        fig = px.line(df_table5, x=df_table5.columns[0], y=yaxis_column_name)
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정

    # 그래프의 배경을 투명하게 설정
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        # plot_bgcolor='rgba(0,0,0,0)',
        yaxis=dict(domain=[0, 0.5]), #하단반쪽   cf)상단반쪽 [0.5,1]

    )

    return fig



@app.callback(
    Output('line-graph-table6', 'figure'),
    [Input('yaxis-column-table6', 'value'),
     Input('graph-type-table6', 'value')]
)
def update_graph_table6(yaxis_column_name, graph_type):
    if graph_type == 'Dot':
        # Dot 그래프 생성
        fig = go.Figure(data=go.Scatter(
            x=df_table6[df_table6.columns[0]].astype(str),
            y=df_table6[yaxis_column_name],
            mode='markers',  # 마커 모드 사용
            marker=dict(
                size=15,  # 데이터 마크 크기 조절 (원하는 크기로 설정)
                opacity=0.8,
                line=dict(width=2, color='DarkSlateGrey')
            )
        ))
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정
        fig.update_xaxes(tickformat="Value", tickangle=-45, tickmode='auto', nticks=12) # 여기서 nticks를 조정하여 간격 변경
        
    elif graph_type == 'Bar':
        fig = px.bar(df_table6, x=df_table6.columns[0], y=yaxis_column_name)
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정
    elif graph_type == 'Pie':
        fig = px.pie(df_table6, names=df_table6.columns[0], values=yaxis_column_name)
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정
    else:
        fig = px.line(df_table6, x=df_table6.columns[0], y=yaxis_column_name)
        fig.update_yaxes(tickformat=".1%")  # Y 축 형식을 백분율로 설정

    # 그래프의 배경을 투명하게 설정
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        # plot_bgcolor='rgba(0,0,0,0)',
        height=700,
    )

    return fig

# 앱 실행
if __name__ == '__main__':
    app.run_server(debug=True,host='0.0.0.0') 


    #http://192.168.194.140:8050